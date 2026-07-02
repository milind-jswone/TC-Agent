from __future__ import annotations

import json
import shutil
import base64
import os
import re
from datetime import UTC, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from tc_parser import TCRecord, parse_supplier_tc


AGENT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = AGENT_ROOT / "Templates" / "tc_output__template.xlsx"
DEFAULT_OUTPUT_DIR = AGENT_ROOT / "Output"
DEFAULT_MASTER = AGENT_ROOT / "Master" / "supplier_tc_master.xlsx"

MAX_TEMPLATE_ROWS = 9
FIRST_ITEM_ROW = 15
IST = timezone(timedelta(hours=5, minutes=30), name="IST")
MASTER_TC_HEADERS = [
    "tc_unique_id", "source_file_name", "source_file_path", "extraction_timestamp", "extraction_model",
    "document_type", "certificate_type", "test_certificate_number", "certificate_date", "document_page_number",
    "total_pages", "document_template_code", "certification_type", "standard_name", "grade", "specification",
    "manufacturer_name", "manufacturing_plant", "plant_location", "district", "state", "country",
    "registered_office", "cin_number", "bis_license_number", "authorized_signatory", "signatory_designation",
    "customer_name", "consignee_name", "consignee_address", "customer_location", "sales_order_number",
    "sales_order_date", "billing_document_number", "invoice_number", "mode_of_transport", "vehicle_number",
    "conforms_to_standard", "conforms_to_is2062", "conforms_to_en10204_type31", "conforms_to_rohs",
    "radioactive_content_compliance", "dimensional_tolerance_compliance", "overall_material_conformance",
    "product_name", "product_category", "material_type", "material_grade", "specification_grade", "steel_type",
    "killing_practice", "manufacturing_process_route", "process_bof", "process_ars", "process_lhf",
    "process_ccm", "process_rh", "process_hsm", "chemical_spec_c_min", "chemical_spec_c_max",
    "chemical_spec_mn_min", "chemical_spec_mn_max", "chemical_spec_s_min", "chemical_spec_s_max",
    "chemical_spec_p_min", "chemical_spec_p_max", "chemical_spec_si_min", "chemical_spec_si_max",
    "chemical_spec_al_min", "chemical_spec_al_max", "chemical_spec_n_min", "chemical_spec_n_max",
    "chemical_spec_b_min", "chemical_spec_b_max", "chemical_spec_nb_min", "chemical_spec_nb_max",
    "chemical_spec_v_min", "chemical_spec_v_max", "chemical_spec_ti_min", "chemical_spec_ti_max",
    "chemical_spec_cr_min", "chemical_spec_cr_max", "chemical_spec_mo_min", "chemical_spec_mo_max",
    "chemical_spec_ni_min", "chemical_spec_ni_max", "chemical_spec_cu_min", "chemical_spec_cu_max",
    "chemical_spec_mae_min", "chemical_spec_mae_max", "chemical_spec_carbon_equivalent_min",
    "chemical_spec_carbon_equivalent_max", "tensile_test_direction_requirement",
    "yield_strength_min_requirement", "yield_strength_max_requirement",
    "ultimate_tensile_strength_min_requirement", "ultimate_tensile_strength_max_requirement",
    "gauge_length_requirement", "elongation_min_requirement", "elongation_max_requirement",
    "yield_to_tensile_ratio_requirement", "bend_test_direction_requirement", "bend_diameter_requirement",
    "bend_test_requirement", "cvn_impact_requirement", "hardness_requirement", "grain_size_requirement",
    "inclusion_rating_requirement", "hole_expansion_ratio_requirement", "erichsen_cupping_requirement",
    "strain_age_embrittlement_requirement", "total_packets", "total_coils", "total_batches", "total_pieces",
    "total_weight_mt", "heat_count", "unique_cast_count", "chemical_composition_pass",
    "mechanical_properties_pass", "bend_test_pass", "dimensional_compliance_pass", "overall_pass_fail",
    "radioactive_compliance_statement", "rohs_compliance_statement", "dimensions_and_tolerance_statement",
    "certification_statement", "remarks", "thickness_unit", "width_unit", "length_unit", "weight_unit",
    "chemistry_unit", "nitrogen_unit", "boron_unit", "yield_strength_unit", "tensile_strength_unit",
    "elongation_unit", "hardness_unit", "impact_energy_unit", "batch_row_number", "cast_number",
    "heat_number", "coil_number", "packet_number", "batch_identifier", "lot_number", "thickness",
    "width", "length", "nominal_size", "pieces_count", "quantity_mt", "carbon_percent", "manganese_percent",
    "sulphur_percent", "phosphorus_percent", "silicon_percent", "aluminium_percent", "nitrogen_ppm",
    "boron_ppm", "niobium_percent", "vanadium_percent", "titanium_percent", "chromium_percent",
    "molybdenum_percent", "nickel_percent", "copper_percent", "micro_alloying_elements_percent",
    "carbon_equivalent_percent", "tensile_test_direction", "yield_strength_mpa",
    "ultimate_tensile_strength_mpa", "gauge_length", "elongation_percent", "yield_to_tensile_ratio",
    "bend_test_direction", "bend_diameter", "bend_test_result", "cvn_impact_direction",
    "cvn_impact_temperature_c", "cvn_impact_average_energy_j", "hardness_hv10", "hardness_hrb",
    "grain_size", "inclusion_rating", "hole_expansion_ratio", "erichsen_cupping_value",
    "strain_age_embrittlement_test",
]


def _safe_name(value: str) -> str:
    safe = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in value.strip())
    return safe.strip("_") or "supplier_tc"


def _now_ist() -> datetime:
    return datetime.now(UTC).astimezone(IST)


def _output_name_for_input(input_path: Path) -> str:
    timestamp = _now_ist().strftime("%Y%m%d_%H%M%S")
    return f"{_safe_name(input_path.stem)}_output_{timestamp}.xlsx"


def _load_record(input_path: str | Path, tc_format: str = "auto") -> TCRecord:
    return parse_supplier_tc(input_path, tc_format=tc_format)


def _set(ws: Any, cell: str, value: Any) -> None:
    target = ws[cell]
    for merged_range in ws.merged_cells.ranges:
        if cell in merged_range:
            target = ws.cell(merged_range.min_row, merged_range.min_col)
            break
    target.value = "" if value is None else value


def _fill_output_template(record: TCRecord, template_path: Path, output_path: Path) -> None:
    if len(record.line_items) > MAX_TEMPLATE_ROWS:
        raise ValueError(
            f"Template has {MAX_TEMPLATE_ROWS} item rows, but {len(record.line_items)} line items were extracted."
        )

    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    _set(ws, "E6", record.test_certificate_no)
    _set(ws, "Y6", record.date)
    _set(ws, "C7", f"To M/S {record.customer_name_address}" if record.customer_name_address else "To M/S")
    _set(ws, "Y7", record.so_no)
    _set(ws, "Y8", record.product)
    _set(ws, "Y9", "")
    _set(ws, "L10", record.test_certificate_no)
    _set(ws, "E11", record.specification)

    for index, item in enumerate(record.line_items):
        row = FIRST_ITEM_ROW + index
        _set(ws, f"C{row}", item.batch_no)
        _set(ws, f"D{row}", record.grade)
        _set(ws, f"E{row}", record.test_certificate_no)
        _set(ws, f"F{row}", item.coil_no)
        _set(ws, f"G{row}", item.nominal_size)
        _set(ws, f"H{row}", item.net_weight_mt)
        _set(ws, f"I{row}", item.width_mm)
        _set(ws, f"J{row}", item.thickness_mm)
        _set(ws, f"K{row}", item.c)
        _set(ws, f"L{row}", item.s)
        _set(ws, f"M{row}", item.p)
        _set(ws, f"N{row}", item.si)
        _set(ws, f"O{row}", item.al)
        _set(ws, f"P{row}", item.n)
        _set(ws, f"Q{row}", item.nb_v_ti)
        _set(ws, f"R{row}", item.tensile_direction)
        _set(ws, f"S{row}", item.tensile_direction)
        _set(ws, f"T{row}", item.ys)
        _set(ws, f"U{row}", item.uts)
        _set(ws, f"V{row}", item.gl)
        _set(ws, f"W{row}", item.elongation)
        _set(ws, f"X{row}", item.bend_direction)
        _set(ws, f"Y{row}", item.bend_radius)
        _set(ws, f"Z{row}", item.bend_result)

    _set(ws, "H24", record.total_weight_mt or sum(item.net_weight_mt for item in record.line_items))
    _set(ws, "P24", record.total_coils_packets or len(record.line_items))
    _set(ws, "E28", record.vehicle_no)
    _set(ws, "E29", record.invoice_no)

    wb.save(output_path)


def _master_headers() -> list[str]:
    return MASTER_TC_HEADERS


def _first_present(*values: Any) -> Any:
    for value in values:
        if value is not None and value != "":
            return value
    return ""


def _payload_value(payload: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in payload and payload[key] not in (None, ""):
            return payload[key]
    return ""


def _next_tc_unique_id(ws: Any) -> str:
    max_id = 0
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        match = re.fullmatch(r"TC_(\d+)", str(value or "").strip(), flags=re.IGNORECASE)
        if match:
            max_id = max(max_id, int(match.group(1)))
    return f"TC_{max_id + 1:03d}"


def _find_or_reset_master_sheet(wb: Any) -> Any:
    for ws in wb.worksheets:
        headers = [ws.cell(row=1, column=index).value for index in range(1, len(MASTER_TC_HEADERS) + 1)]
        if headers == MASTER_TC_HEADERS:
            return ws

    for ws in wb.worksheets:
        if ws.cell(row=1, column=1).value == "tc_unique_id":
            for index, header in enumerate(MASTER_TC_HEADERS, start=1):
                ws.cell(row=1, column=index).value = header
            return ws

    while wb.worksheets:
        wb.remove(wb.worksheets[0])
    ws = wb.create_sheet("TC Data")
    ws.append(MASTER_TC_HEADERS)
    return ws


def _record_common_values(record: TCRecord, processed_at: str) -> dict[str, Any]:
    raw = record.raw_data or {}
    customer_address = record.customer_name_address
    return {
        "source_file_name": record.source_file,
        "source_file_path": record.source_file,
        "extraction_timestamp": processed_at,
        "extraction_model": os.getenv("LLM_MODEL", "gemini-2.5-flash"),
        "document_type": _payload_value(raw, "document_type") or "TEST CERTIFICATE",
        "certificate_type": _payload_value(raw, "certificate_type") or "Test Certificate",
        "test_certificate_number": record.test_certificate_no,
        "certificate_date": record.date,
        "grade": record.grade,
        "specification": record.specification,
        "customer_name": _payload_value(raw, "customer_name", "consignee_name") or customer_address,
        "consignee_name": _payload_value(raw, "consignee_name", "customer_name") or customer_address,
        "consignee_address": _payload_value(raw, "consignee_address", "customer_location") or customer_address,
        "customer_location": _payload_value(raw, "customer_location", "consignee_address") or customer_address,
        "sales_order_number": record.so_no,
        "sales_order_date": record.so_date,
        "billing_document_number": record.billing_doc_no,
        "invoice_number": record.invoice_no,
        "vehicle_number": record.vehicle_no,
        "product_name": record.product,
        "product_category": record.product,
        "material_grade": record.grade,
        "specification_grade": record.specification,
        "standard_name": _payload_value(raw, "standard_name") or (record.specification.split()[0] if record.specification else ""),
        "total_packets": record.total_coils_packets,
        "total_coils": record.total_coils_packets,
        "total_batches": record.total_coils_packets,
        "total_weight_mt": record.total_weight_mt,
        "thickness_unit": "mm",
        "width_unit": "mm",
        "length_unit": "mm",
        "weight_unit": "MT",
        "chemistry_unit": "%",
        "nitrogen_unit": "ppm",
        "boron_unit": "ppm",
        "yield_strength_unit": "MPa",
        "tensile_strength_unit": "MPa",
        "elongation_unit": "%",
    }


def _line_item_values(item: Any, row_number: int) -> dict[str, Any]:
    raw = item.raw_data or {}
    return {
        "batch_row_number": row_number,
        "cast_number": _payload_value(raw, "cast_number") or item.batch_no,
        "heat_number": _payload_value(raw, "heat_number") or item.batch_no,
        "coil_number": _payload_value(raw, "coil_number") or item.coil_no,
        "packet_number": _payload_value(raw, "packet_number") or item.coil_no,
        "batch_identifier": _payload_value(raw, "batch_identifier") or item.batch_no,
        "thickness": _payload_value(raw, "thickness") or item.thickness_mm,
        "width": _payload_value(raw, "width") or item.width_mm,
        "length": _payload_value(raw, "length") or item.length_mm,
        "nominal_size": item.nominal_size,
        "pieces_count": _payload_value(raw, "pieces_count") or item.pcs,
        "quantity_mt": _payload_value(raw, "quantity_mt") or item.net_weight_mt,
        "carbon_percent": _payload_value(raw, "carbon_percent") or item.c,
        "sulphur_percent": _payload_value(raw, "sulphur_percent") or item.s,
        "phosphorus_percent": _payload_value(raw, "phosphorus_percent") or item.p,
        "silicon_percent": _payload_value(raw, "silicon_percent") or item.si,
        "aluminium_percent": _payload_value(raw, "aluminium_percent") or item.al,
        "nitrogen_ppm": _payload_value(raw, "nitrogen_ppm") or item.n,
        "niobium_percent": _payload_value(raw, "niobium_percent") or item.nb,
        "vanadium_percent": _payload_value(raw, "vanadium_percent") or item.v,
        "titanium_percent": _payload_value(raw, "titanium_percent") or item.ti,
        "micro_alloying_elements_percent": _payload_value(raw, "micro_alloying_elements_percent") or item.nb_v_ti,
        "tensile_test_direction": _payload_value(raw, "tensile_test_direction") or item.tensile_direction,
        "yield_strength_mpa": _payload_value(raw, "yield_strength_mpa") or item.ys,
        "ultimate_tensile_strength_mpa": _payload_value(raw, "ultimate_tensile_strength_mpa") or item.uts,
        "gauge_length": _payload_value(raw, "gauge_length") or item.gl,
        "elongation_percent": _payload_value(raw, "elongation_percent") or item.elongation,
        "yield_to_tensile_ratio": _payload_value(raw, "yield_to_tensile_ratio") or item.ys_uts_ratio,
        "bend_test_direction": _payload_value(raw, "bend_test_direction") or item.bend_direction,
        "bend_diameter": _payload_value(raw, "bend_diameter") or item.bend_radius,
        "bend_test_result": _payload_value(raw, "bend_test_result") or item.bend_result,
    }


def _master_row(record: TCRecord, item: Any, tc_unique_id: str, processed_at: str, row_number: int) -> list[Any]:
    raw = record.raw_data or {}
    item_raw = item.raw_data or {}
    common = _record_common_values(record, processed_at)
    item_values = _line_item_values(item, row_number)
    row = []
    for header in MASTER_TC_HEADERS:
        row.append(
            _first_present(
                tc_unique_id if header == "tc_unique_id" else "",
                item_raw.get(header),
                raw.get(header),
                item_values.get(header),
                common.get(header),
            )
        )
    return row


def _append_master(record: TCRecord, master_path: Path) -> None:
    master_path.parent.mkdir(parents=True, exist_ok=True)
    headers = _master_headers()
    if master_path.exists():
        wb = openpyxl.load_workbook(master_path)
        ws = _find_or_reset_master_sheet(wb)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TC Data"
        ws.append(headers)

    processed_at = _now_ist().strftime("%Y-%m-%d %H:%M:%S IST")
    tc_unique_id = _next_tc_unique_id(ws)
    for row_number, item in enumerate(record.line_items, start=1):
        ws.append(_master_row(record, item, tc_unique_id, processed_at, row_number))

    if "SupplierTCMaster" not in ws.tables and ws.max_row > 1:
        table_ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        table = Table(displayName="SupplierTCMaster", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
    elif "SupplierTCMaster" in ws.tables:
        ws.tables["SupplierTCMaster"].ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    for column in ws.columns:
        letter = column[0].column_letter
        max_len = max(len(str(cell.value or "")) for cell in column[:50])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)
    wb.save(master_path)


def process_supplier_tc(
    input_path: str | Path,
    template_path: str | Path = DEFAULT_TEMPLATE,
    output_dir: str | Path = DEFAULT_OUTPUT_DIR,
    master_path: str | Path = DEFAULT_MASTER,
    include_file_content: bool = False,
    include_master_file_content: bool = False,
    tc_format: str = "auto",
) -> dict[str, Any]:
    input_path = Path(input_path)
    template_path = Path(template_path)
    output_dir = Path(output_dir)
    master_path = Path(master_path)
    output_dir.mkdir(parents=True, exist_ok=True)

    record = _load_record(input_path, tc_format=tc_format)
    output_name = _output_name_for_input(input_path)
    output_path = output_dir / output_name
    json_path = output_dir / f"{output_path.stem}.json"

    _fill_output_template(record, template_path, output_path)
    _append_master(record, master_path)
    json_path.write_text(json.dumps(record.to_dict(), indent=2), encoding="utf-8")

    result = {
        "status": "success" if not record.warnings else "success_with_warnings",
        "source_file": str(input_path),
        "output_file_name": output_path.name,
        "output_file": str(output_path),
        "master_file": str(master_path),
        "json_file": str(json_path),
        "line_items": len(record.line_items),
        "tc_format": tc_format,
        "extracted_data": record.to_dict(),
        "warnings": record.warnings,
    }
    if include_file_content:
        result["output_file_base64"] = base64.b64encode(output_path.read_bytes()).decode("ascii")
    if include_master_file_content:
        result["master_file_name"] = master_path.name
        result["master_file_base64"] = base64.b64encode(master_path.read_bytes()).decode("ascii")
    return result


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="Generate JSW One TC output from supplier TC PDF or image.")
    parser.add_argument("input_path")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--master", default=str(DEFAULT_MASTER))
    parser.add_argument("--tc-format", default="auto")
    args = parser.parse_args()

    result = process_supplier_tc(args.input_path, args.template, args.output_dir, args.master, tc_format=args.tc_format)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
